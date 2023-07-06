import copy
import json

import bson
import pymongo
import tempfile
import datetime
from bson import ObjectId
from fastapi import APIRouter, HTTPException, status, Depends, BackgroundTasks, Request, UploadFile

from backend.services.excel import ExcelService
from backend.services.logs import LogsService
from backend.services.mongo import MongoService
from backend.db.base import Data, Modules, Config, Adjustment, AssetExtraData
from backend.models.assets import AssetModel, ForceUpdateImport
from backend.services.modules import (
    caluclate_modules_result, check_data_asset, prepare_modules_results
)
from backend.services.oauth2 import check_admin, check_auth
from backend.services.scraping import Scraping
from backend.services.utils import remove_nulls_from_dict, get_fields_different, check_period_asset_valid, \
    find_key_by_value, prepare_data_to_db, update_asset_from_file
from fastapi.responses import FileResponse
from openpyxl import load_workbook

from backend.settings.config import settings
from backend.services.final_model import FinalModel
from pymongo import UpdateOne

router = APIRouter()
AVAILABLE_ASSETS = settings.AVAILABLE_ASSETS
AVAILABLE_PERIOD = settings.AVAILABLE_PERIOD


@router.get("/get_all_assets")
def get_all_last(
        auth_user: dict = Depends(check_admin)
):
    all_assets_extra_data = AssetExtraData.find({})
    all_adjustment = Adjustment.find({})
    adjustment = copy.deepcopy(list(all_adjustment))
    assets_extra_data = copy.deepcopy(list(all_assets_extra_data))
    adjustment_map = {}
    extra_data_map = {elem.get('type_asset'): MongoService.data_to_json(elem) for elem in assets_extra_data}

    for adj in adjustment:
        for asset in adj['assets']:
            asset_id = str(adj.get('_id'))
            if asset in adjustment_map:
                adjustment_map[asset].backendend(asset_id)
            else:
                adjustment_map[asset] = [asset_id]

    all_assets = Data.aggregate(
        [
            {
                "$group": {
                    "_id": {"period": "$period", "type_asset": "$type_asset"},
                    "data": {
                        "$last": {
                            "_id": "$_id",
                            **MongoService.mbackended_keys_mongo(
                                AssetModel().get_fields_name()
                            ),
                        }
                    },
                },
            },
            {"$sort": {"data.date": -1}},
            {
                "$group": {
                    "_id": "$_id.type_asset",
                    "values": {"$push": {"key": "$_id.period", "data": "$data"}},
                }
            },
        ]
    )
    result_dict = {}
    for asset in all_assets:
        result_dict[asset["_id"]] = {
            'adjustments': adjustment_map.get(asset['_id'], []),
            'extra_data': extra_data_map.get(asset['_id'], {})
        }
        for period_record in asset["values"]:
            result_dict[asset["_id"]][period_record["key"]] = [
                MongoService.data_to_json(period_record["data"])
            ]
    return result_dict


def check_available_adjustment(adjustments, asset):
    type_asset = asset.get('type_asset')
    for adj in list(adjustments):
        if type_asset not in adj.get('assets'):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                                detail=f'"{adj.get("name")}" Adjustment not supported {type_asset}')


@router.post("/asset/{asset_id}", status_code=status.HTTP_200_OK)
def asset_update(payload: AssetModel, asset_id,
                 background_tasks: BackgroundTasks,
                 auth_user: dict = Depends(check_admin),
                 user: dict = Depends(check_auth),
                 ):
    user_id = user['id']
    MongoService.is_valid_mongo_id(asset_id, error_msg='Asset id is invalid')
    dict_payload = remove_nulls_from_dict(payload.dict())
    asset = Data.find_one({'_id': ObjectId(asset_id)})
    if not asset:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f'Asset with id {asset_id} not found.')
    data_for_update = MongoService.convert_2_dict_mongodb(dict_payload, deleted_keys=['date'])
    result = Data.find_one_and_update(
        {"_id": ObjectId(str(asset_id))}, {"$set": data_for_update}
    )
    if not result:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail='Smth went wrong')

    try:
        background_tasks.add_task(LogsService,
                                  before_object=MongoService.convert_2_dict_mongodb(asset, deleted_keys=['date']),
                                  after_object=data_for_update, user_id=user_id, type_='asset', action='update',
                                  fields=AssetModel().get_fields_name(fields='__for_changes__'),
                                  obj={'id': str(asset.get('_id')),
                                       'type_asset': asset.get('type_asset'),
                                       'period': asset.get('period'),
                                       'date': asset.get('date')})
    except ValueError as e:
        print("Error create logs", e)
    return {
        'result': True
    }


@router.get("/asset/{asset_id}", status_code=status.HTTP_200_OK)
def get_one(asset_id,
            auth_user: dict = Depends(check_admin),
            user: dict = Depends(check_auth),
            ):
    try:
        ObjectId(str(asset_id))
    except bson.errors.InvalidId as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid object ID",
        )
    result = Data.find_one({"_id": ObjectId(str(asset_id))})
    return MongoService.data_to_json(result)


@router.get("/get_assets")
def get_data(
        type_asset: str,
        period: str,
        page: int = 1,
        count: str = 30,
        date_from=None,
        date_to=None,
        auth_user: dict = Depends(check_admin),
        user: dict = Depends(check_auth),
):
    check_period_asset_valid(period=period, type_asset=type_asset)

    date_from_db, date_to_db = None, None
    if date_from:
        date_from_db = MongoService.string_to_date_mongo(date_from)

    if date_to:
        date_to_db = MongoService.string_to_date_mongo(date_to)

    filter_fields = {
        "period": period,
        "type_asset": type_asset,
    }

    if date_from_db and date_to_db:
        filter_fields["date"] = {"$gte": date_from_db, "$lt": date_to_db}
    if count == "all":
        result = Data.find(filter_fields).sort([("date", pymongo.DESCENDING)])
    else:
        count = int(count)
        skips = count * (page - 1)
        result = (
            Data.find(filter_fields)
            .sort([("date", pymongo.DESCENDING)])
            .skip(skips)
            .limit(count)
        )
    extra_data = AssetExtraData.find_one({'type_asset': type_asset})
    return {
        "data": MongoService.data_list_to_json(result),
        "extra_data": MongoService.data_to_json(extra_data),
        "count": Data.count_documents(filter_fields),
    }


@router.get('/get_modules/check_empty_fields/{asset_id}')
def check_module(request: Request, asset_id: str,
                 auth_user: dict = Depends(check_admin),
                 user: dict = Depends(check_auth),
                 ):
    asset = dict(Data.find_one({"_id": ObjectId(asset_id)}))
    result_check_fields = check_data_asset(asset)
    return {'result': result_check_fields}


@router.get("/get_modules/{asset_id}", status_code=status.HTTP_200_OK)
def modules_result(request: Request, asset_id: str,
                   auth_user: dict = Depends(check_admin),
                   user: dict = Depends(check_auth),
                   ):
    try:
        ObjectId(str(asset_id))
    except bson.errors.InvalidId as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid object ID",
        )
    modules = Modules.find({}).sort([("sort_key", 1)])
    asset = dict(Data.find_one({"_id": ObjectId(asset_id)}))
    adjustments = list(Adjustment.find({"assets": {"$in": [asset['type_asset']]}}))
    list_modules = list(modules)

    result = caluclate_modules_result(list_modules, asset, adjustments=adjustments)

    for adj_id, modules_result in result.items():
        if len(modules_result) > 3:
            final_model = FinalModel(
                adjustments, result[adj_id], asset, adj_id, list_modules
            )
            try:
                final_decision_module = final_model.get_result()
                result[adj_id].update(final_decision_module)
            except Exception as e:

                result[adj_id]["Module P"] = str(e)

    return {"result": prepare_modules_results(result)}


@router.get("/download")
def download_excel(
        type_asset: str,
        period: str,
        date_from=None,
        date_to=None,
        adjustment_id=None,
        auth_user: dict = Depends(check_admin),
        user: dict = Depends(check_auth),
):
    check_period_asset_valid(period=period, type_asset=type_asset)

    workbook = load_workbook('template_file.xlsx')

    date_from_db, date_to_db = None, None
    if date_from:
        date_from_db = MongoService.string_to_date_mongo(date_from)

    if date_to:
        date_to_db = MongoService.string_to_date_mongo(date_to)

    filter_fields = {"period": period, "type_asset": type_asset}

    if date_from_db and date_to_db:
        filter_fields["date"] = {"$gte": date_from_db, "$lt": date_to_db}

    result = Data.find(filter_fields).sort([("date", pymongo.DESCENDING)])

    filter_fields_adj = {"_id": ObjectId(adjustment_id)} if adjustment_id else {"name": "basic"}

    adjustment = Adjustment.find_one(filter_fields_adj)

    adjustment_id = str(adjustment.get('_id'))
    adjustment_name = adjustment.get("name")

    buy_sell_percent = adjustment.get('buy_sell_percent')
    buy_plus = buy_sell_percent.get('buy_plus')
    sell_minus = buy_sell_percent.get('sell_minus')

    trend_values_1 = adjustment.get('trend_values', {}).get('1')
    trend_values_2 = adjustment.get('trend_values', {}).get('2')

    sheet = workbook.active
    map_period = {'d': 'Daily', 'm': 'Monthly', 'w': 'Weekly'}
    sheet['A1'] = f'Adjustment: {adjustment_name}'
    sheet['B1'] = f'Asset: {type_asset}| Period {map_period[period]}'
    sheet['C1'] = f'Date from: {date_from}'
    sheet['D1'] = f'Date to: {date_to}'

    start_row_number = 4
    sheet['H3'] = f'sell - {sell_minus}%'
    sheet['I3'] = f'buy + {buy_plus}%'
    sheet['N3'] = trend_values_1
    sheet['O3'] = trend_values_2

    max_empty = 0

    for index, asset in enumerate(result):
        buy_sell = asset.get('buy_sell', [])
        levels_crossed = asset.get('levels_crossed', [])
        trend_values = asset.get('trend_values', {})

        trend_values_0 = trend_values.get('0', )
        trend_values_3 = trend_values.get('3', )

        len_buy_sell = len(buy_sell)
        len_levels_crossed = len(levels_crossed)
        max_len = max(len_buy_sell, len_levels_crossed)
        adjustment_ids = asset.get('adjustment_ids', {}).get(adjustment_id, {})

        col_row_index = start_row_number + max_empty

        sheet[f'A{col_row_index}'] = asset.get('date')
        open = adjustment_ids.get('open')
        low = adjustment_ids.get('low')
        high = adjustment_ids.get('high')
        close = adjustment_ids.get('close')

        sheet[f'B{col_row_index}'] = open if open else asset.get('open')
        sheet[f'C{col_row_index}'] = low if low else asset.get('low')
        sheet[f'D{col_row_index}'] = high if high else asset.get('high')
        sheet[f'E{col_row_index}'] = close if close else asset.get('close')

        sheet[f'L{col_row_index}'] = asset.get('trend_indicator')
        sheet[f'M{col_row_index}'] = trend_values_0

        if trend_values_3 and trend_values_0:
            sheet[f'N{col_row_index}'] = (trend_values_3 - trend_values_0) * trend_values_1 / 100 + trend_values_0
            sheet[f'O{col_row_index}'] = (trend_values_3 - trend_values_0) * trend_values_2 / 100 + trend_values_0

        sheet[f'Q{col_row_index}'] = asset.get('trend_length')
        sheet[f'R{col_row_index}'] = asset.get('trend_percent')
        sheet[f'S{col_row_index}'] = asset.get('momentum')
        sheet[f'T{col_row_index}'] = asset.get('timing')

        sheet[f'P{col_row_index}'] = trend_values_3

        sorted_buy_sell = sorted(buy_sell, key=lambda x: x.get('sort_key'))
        sorted_levels_crossed = sorted(levels_crossed, key=lambda x: x.get('sort_key'))

        for index_sell, buy_sell_row in enumerate(sorted_buy_sell):
            row_buy_sell_number = start_row_number + index_sell + max_empty
            sell_value = buy_sell_row.get('sell') or 0
            buy_value = buy_sell_row.get('buy') or 0
            sheet[f'F{row_buy_sell_number}'] = sell_value if sell_value else ''
            sheet[f'G{row_buy_sell_number}'] = buy_value if buy_value else ''

            sheet[f'H{row_buy_sell_number}'] = sell_value * ((100 - sell_minus) / 100) if sell_value else ''
            sheet[f'I{row_buy_sell_number}'] = buy_value * ((100 + buy_plus) / 100) if buy_value else ''

        for index_levels, level_crossed_row in enumerate(sorted_levels_crossed):
            row_levels_number = start_row_number + index_levels + max_empty
            levels_sell = level_crossed_row.get('sell')
            levels_buy = level_crossed_row.get('buy')

            sheet[f'J{row_levels_number}'] = levels_sell
            sheet[f'K{row_levels_number}'] = levels_buy

        max_empty += max_len if max_len else 1

    # Save the workbook to a temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        file_path = tmp.name
        workbook.save(file_path)
        now = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        file_name = f'{type_asset}_{adjustment_name}_{now}.xlsx'
        return FileResponse(file_path, filename=file_name,
                            media_type="backendlication/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def fill_data_db():
    scraping = Scraping()
    scraping.fill_data()


@router.get("/fill_data")
def fill_data(
        auth_user: dict = Depends(check_admin),
        user: dict = Depends(check_auth),
):
    fill_data_db()


@router.get("/clear_update_db")
def clear_update_db(
        auth_user: dict = Depends(check_admin),
        user: dict = Depends(check_auth),
):
    Config.delete_many({"type": "schedule"})
    Data.delete_many({})
    fill_data_db()


@router.get("/clear_duplicate")
def clear_duplicate(
        auth_user: dict = Depends(check_admin),
        user: dict = Depends(check_auth),
):
    delete_ids = []
    first_asset_date = None
    for asset in Data.find().sort([("type_asset", -1), ("period", -1), ("date", -1), ("trend_values", -1)]):
        if asset.get("date") != first_asset_date:
            first_asset_date = asset.get("date")
            continue
        delete_ids.backendend(asset.get("_id"))
    Data.delete_many({"_id": {"$in": delete_ids}})


@router.post("/import_excel")
async def create_upload_file(file: UploadFile,
                             auth_user: dict = Depends(check_admin),
                             user: dict = Depends(check_auth),
                             ):
    try:
        excel_service = ExcelService(file)
        sheet = excel_service.get_active_sheet()
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST , detail=str(e))

    type_asset = sheet['A1'].value
    period = sheet['B1'].value

    map_period = {'daily': 'd', 'weekly': 'w', 'monthly': 'm'}

    if not type_asset:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Column A1 should contain type of asset")
    if not period:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Column B1 should contain type of period {','.join(list(map_period.keys()))}")

    type_asset = type_asset.upper().strip()
    period = period.lower().strip()

    period = map_period[period]
    max_row, max_col = excel_service.get_last_row_position(sheet)

    if not max_row or not max_col:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail='File can not to be empty.')
    map_col = {
        'date': 0,
        'open': 1,
        'low': 2,
        'high': 3,
        'close': 4,
        'buy_sell.sell': 5,
        'buy_sell.buy': 6,
        'levels_crossed.sell': 9,
        'levels_crossed.buy': 10,
        'trend_indicator': 11,
        'trend_values.0': 12,
        'trend_values.3': 15,
        'trend_length': 16,
        'trend_percent': 17,
        'momentum': 18,
        'timing': 19,
    }

    keys = list(map_col.keys())
    values = list(map_col.values())
    result_dict = {}
    target_date = None

    for index_row, row in enumerate(sheet.iter_rows(min_row=4, min_col=1, max_col=max_col, max_row=max_row)):
        date_as_key = row[0].value
        if date_as_key not in result_dict and date_as_key:
            date_as_key = date_as_key
            result_dict[date_as_key] = {}
            target_date = date_as_key
        for index_col, col in enumerate(values):
            row_value = row[col].value
            if col == 18 and row_value:  # momentum
                row_value = row_value.replace(' ', '').upper()

            if isinstance(row_value, str):
                row_value = row_value.replace(' ', '') or None  # '' or None -> None
            try:
                position_in_keys = values.index(col)
                target_key = keys[position_in_keys]  # open , close , etc
                if target_key != 'date':
                    if target_key not in result_dict[target_date]:
                        result_dict[target_date][target_key] = [row_value]
                    else:
                        result_dict[target_date][target_key].backendend(row_value)
            except ValueError as e:
                print(e)

    clear_data_to_db = prepare_data_to_db(result_dict)

    # just for clarity testing, create result.json file with content "[ ]"  in fuzzy_back folder
    # with open('result.json', 'w') as f:
    #     json.dump(clear_data_to_db, f)

    result_update = update_asset_from_file(type_asset=type_asset, period=period, data=clear_data_to_db)
    returned_object = {'result': True}
    returned_object.update(result_update)
    return returned_object


@router.post("/force_update")
def force_update_record_after_import(payload: ForceUpdateImport,
                                     auth_user: dict = Depends(check_admin),
                                     user: dict = Depends(check_auth),
                                     ):
    dict_payload = payload.dict()
    records = dict_payload.get('records')
    if not records:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f'"No records.')

    type_asset = dict_payload.get('type_asset')
    period = dict_payload.get('period')
    check_period_asset_valid(period=period, type_asset=type_asset)

    dates = list(map(lambda date: MongoService.string_to_date_mongo(date, "%Y-%m-%d %H:%M:%S"), list(records.keys())))
    assets = list(Data.find({'date': {"$in": dates}, 'type_asset': "AUD", 'period': 'd'}))
    pipeline = []

    for asset in assets:
        pipeline.backendend(
            UpdateOne({"_id": asset.get('_id')}, {'$set': records[str(asset['date'])]})
        )
    result = Data.bulk_write(pipeline)
    returned_object = {'result': True}
    returned_object.update(
        {
            'matched_count': result.matched_count,
            'inserted_count': result.inserted_count,
            'modified_count': result.modified_count,
            'deleted_count': result.deleted_count,
            'upserted_ids': result.upserted_ids,
        }
    )
    return returned_object
