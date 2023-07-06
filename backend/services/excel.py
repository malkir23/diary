from openpyxl import load_workbook


class ExcelService:
    def __init__(self, file):
        self.exts = (".xlsx", ".xls")
        self.file = file
        self.filename = file.filename
        """
             65 - ord("A") , 66 - ord("B") , .... , 85 - ord("T")
        """
        self.VALID_COL_NAME_MAP_TO_IMPORT = {65: 'date', 66: 'open', 67: 'low', 68: 'high',
                                             69: 'close', 70: 'buy&selllevels', 71: None,
                                             72: 'buy&selllevels', 73: None, 74: 'levelscrossed',
                                             75: None, 76: 'trendindicator', 77: 'trendvalues',
                                             78: None, 79: None, 80: None, 81: 'trendlength',
                                             82: 'trend%', 83: 'momentum', 84: 'timing'}

        self.validate()

    def get_active_sheet(self):
        workbook = load_workbook(self.file.file)
        sheet = workbook.active
        range_letters = range(ord('A'), ord('T') + 1)
        for ord_number in range_letters:
            default_col_name = sheet[f'{chr(ord_number)}2'].value
            name_col = default_col_name.strip().lower().replace(' ', '') if default_col_name else default_col_name
            valid_name_col = self.VALID_COL_NAME_MAP_TO_IMPORT[ord_number]
            if name_col != valid_name_col:
                raise ValueError(f'Invalid column name. Col {chr(ord_number)}2 should be "{valid_name_col}"')
        return sheet

    def get_last_row_position(self, sheet):
        max_row = sheet.max_row
        # Starting from the last row, iterate backwards until a non-empty cell is found
        for row in range(max_row, 0, -1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value is not None:
                last_row = row
                last_column = sheet.max_column
                return last_row, last_column
        return None, None

    def is_excel_file(self) -> bool:
        return self.filename.endswith((".xlsx", ".xls"))

    def validate(self):
        is_excel_file = self.is_excel_file()
        if not is_excel_file:
            raise ValueError(f'File should be excel - ({",".join(self.exts)})')
